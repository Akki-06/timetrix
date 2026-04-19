import math
import logging

from django.db import transaction
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend

import os
from pathlib import Path
from django.http import FileResponse, Http404
from django.conf import settings
log = logging.getLogger(__name__)
from .models import (
    Department,
    Program,
    AcademicTerm,
    Course,
    CourseType,
    StudentGroup,
    CourseOffering
)
from .serializers import (
    DepartmentSerializer,
    ProgramSerializer,
    AcademicTermSerializer,
    CourseSerializer,
    StudentGroupSerializer,
    CourseOfferingSerializer
)

class TemplateDownloadView(APIView):
    """
    Serves standard Excel templates for bulk uploads.
    """
    def get(self, request, *args, **kwargs):
        entity = request.query_params.get("entity", "").lower()
        mapping = {
            "programs": "TIMETRIX_Programs_Template.xlsx",
            "courses" : "TIMETRIX_Courses_Template.xlsx",
            "faculty" : "TIMETRIX_Faculty_Template.xlsx",
            "rooms"   : "TIMETRIX_Rooms_Template.xlsx",
            "offerings": "TIMETRIX_CourseOfferings_Template.xlsx",
        }
        
        filename = mapping.get(entity)
        if not filename:
            return Response({"error": "Invalid entity type"}, status=status.HTTP_400_BAD_REQUEST)
            
        path = Path(settings.BASE_DIR) / "ml_pipeline" / "data" / "templates" / filename
        if not path.exists():
            raise Http404("Template file not found.")
            
        return FileResponse(open(path, 'rb'), as_attachment=True, filename=filename)


# ----------------------------
# DEPARTMENT
# ----------------------------

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "code"]
    ordering_fields = ["name", "code"]


# ----------------------------
# PROGRAM
# ----------------------------

class ProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.select_related("department")
    serializer_class = ProgramSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["department"]
    search_fields = ["name", "code"]


# ----------------------------
# ACADEMIC TERM
# ----------------------------

class AcademicTermViewSet(viewsets.ModelViewSet):
    queryset = AcademicTerm.objects.select_related("program")
    serializer_class = AcademicTermSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["program", "year", "semester"]


# ----------------------------
# COURSE
# ----------------------------

class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related("program").all()
    serializer_class = CourseSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["course_type", "program", "semester"]
    search_fields = ["name", "code"]


# ----------------------------
# STUDENT GROUP
# ----------------------------

class StudentGroupViewSet(viewsets.ModelViewSet):
    queryset = StudentGroup.objects.select_related("term", "term__program", "term__program__department")
    serializer_class = StudentGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["term", "term__program", "term__semester"]
    search_fields = ["name", "description"]

    @action(detail=False, methods=["post"], url_path="quick-create")
    def quick_create(self, request):
        """
        Single-step creation: admin provides program + semester + section + strength.
        Automatically finds or creates the AcademicTerm, then creates the StudentGroup.
        """
        program_id = request.data.get("program")
        semester   = request.data.get("semester")
        section    = request.data.get("section", "").strip()
        strength   = request.data.get("strength")
        description = request.data.get("description", "")

        if not all([program_id, semester, section, strength]):
            return Response(
                {"error": "program, semester, section, and strength are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            program_id = int(program_id)
            semester   = int(semester)
            strength   = int(strength)
        except (ValueError, TypeError):
            return Response({"error": "Invalid numeric values."}, status=status.HTTP_400_BAD_REQUEST)

        if strength <= 0:
            return Response({"error": "Strength must be > 0."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate program exists (404 if not)
        program = get_object_or_404(Program, pk=program_id)

        # Derive academic year from semester (sem 1-2 → year 1, sem 3-4 → year 2 …)
        year = math.ceil(semester / 2)

        # Auto-create AcademicTerm if it doesn't exist yet
        term, _ = AcademicTerm.objects.get_or_create(
            program_id=program.id,
            semester=semester,
            defaults={"year": year},
        )

        working_days = request.data.get("working_days", ["MON", "TUE", "WED", "THU", "FRI"])

        # Create section (or update strength if it already exists)
        group, created = StudentGroup.objects.get_or_create(
            term=term,
            name=section,
            defaults={"strength": strength, "description": description, "working_days": working_days},
        )
        if not created:
            group.strength = strength
            if description is not None:
                group.description = description
            group.working_days = working_days
            group.save()

        data = StudentGroupSerializer(group).data
        return Response(data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="auto-assign-courses")
    def auto_assign_courses(self, request, pk=None):
        """
        Auto-create CourseOfferings for all courses that match this
        section's program + semester.

        POST /api/academics/student-groups/{id}/auto-assign-courses/

        Optional body: {"overwrite": true} to delete existing offerings first.
        """
        group = self.get_object()
        term = group.term
        program = term.program
        semester = term.semester

        overwrite = request.data.get("overwrite", False)

        NON_SCHED = {"DIS", "INT", "RND"}

        all_courses = list(Course.objects.filter(
            program=program,
            semester=semester,
        ).exclude(course_type__in=NON_SCHED))

        if not all_courses:
            return Response(
                {"error": f"No schedulable courses found for {program.code} Sem {semester}."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Collect placeholder codes so we skip assigning the placeholder when actual choices exist
        placeholder_codes = set()
        for c in all_courses:
            if c.parent_course_code:
                placeholder_codes.add(c.parent_course_code)

        courses_to_assign = [c for c in all_courses if c.code not in placeholder_codes]

        if overwrite:
            CourseOffering.objects.filter(student_group=group).delete()

        created_count = 0
        already_existed = 0
        course_codes = []

        for course in courses_to_assign:
            esg = course.parent_course_code if course.parent_course_code else None
            
            offering, was_created = CourseOffering.objects.get_or_create(
                course=course,
                student_group=group,
                defaults={
                    "weekly_load": course.min_weekly_lectures,
                    "elective_slot_group": esg,
                },
            )
            
            # If existed, ensure elective_slot_group is up to date
            if not was_created and offering.elective_slot_group != esg:
                offering.elective_slot_group = esg
                offering.save(update_fields=["elective_slot_group"])
                
            if was_created:
                created_count += 1
            else:
                already_existed += 1
            course_codes.append(course.code)

        return Response({
            "created": created_count,
            "already_existed": already_existed,
            "courses": course_codes,
        })


# ----------------------------
# COURSE OFFERING
# ----------------------------

class CourseOfferingViewSet(viewsets.ModelViewSet):
    queryset = CourseOffering.objects.select_related(
        "course",
        "student_group",
        "assigned_faculty"
    )
    serializer_class = CourseOfferingSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = [
        "course",
        "student_group",
        "assigned_faculty",
        # Related traversals used by the Generator page PE UI
        "course__course_type",
        "student_group__term__program",
        "student_group__term__semester",
    ]


# ----------------------------
# COURSE BULK UPLOAD
# ----------------------------

# Map full names → short codes for course_type
_COURSE_TYPE_MAP = {
    "program core": "PC", "pc": "PC",
    "program elective": "PE", "pe": "PE",
    "open elective": "OE", "oe": "OE",
    "basic sciences course": "BSC", "bsc": "BSC",
    "engineering sciences course": "ESC", "esc": "ESC",
    "humanities": "HUM", "hum": "HUM",
    "life skills": "LS", "ls": "LS",
    "value added module": "VAM", "vam": "VAM",
    "ability enhancement course": "AEC", "aec": "AEC",
    "practical (lab)": "PR", "practical": "PR", "lab": "PR", "pr": "PR",
    "project": "PRJ", "prj": "PRJ",
    "dissertation": "DIS", "dis": "DIS",
    "internship": "INT", "int": "INT",
    "research": "RND", "rnd": "RND",
}

VALID_COURSE_TYPES = {c[0] for c in CourseType.choices}


class CourseBulkUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided. Upload an .xlsx file with field name 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith(".xlsx"):
            return Response(
                {"error": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            import openpyxl
        except ImportError:
            return Response(
                {"error": "openpyxl is not installed. Run: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"error": f"Could not read Excel file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response(
                {"error": "File must have a header row and at least one data row."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

        required = {"code", "name", "credits", "course_type"}
        if not required.issubset(set(header)):
            return Response(
                {"error": f"Missing required columns: {required - set(header)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        col = {name: idx for idx, name in enumerate(header)}
        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    code = str(row[col["code"]] or "").strip()
                    # Normalization: strip _BCA, _BTech etc suffixes and uppercase
                    code = code.split('_')[0].upper()
                    
                    name = str(row[col["name"]] or "").strip()
                    credits_val = int(row[col["credits"]] or 3)

                    if not code or not name:
                        errors.append(f"Row {row_num}: code or name is empty, skipped.")
                        continue

                    # Resolve course_type
                    ct_raw = str(row[col["course_type"]] or "PC").strip()
                    ct = _COURSE_TYPE_MAP.get(ct_raw.lower(), ct_raw.upper())
                    if ct not in VALID_COURSE_TYPES:
                        errors.append(f"Row {row_num}: unknown course_type '{ct_raw}', skipped.")
                        continue

                    # Resolve program (optional) — auto-create if missing
                    program = None
                    semester = None
                    parent_code = None
                    
                    if "program_code" in col and row[col["program_code"]]:
                        prog_code = str(row[col["program_code"]]).strip()
                        program = Program.objects.filter(code__iexact=prog_code).first()
                        if not program:
                            errors.append(f"Row {row_num}: Program code '{prog_code}' not found. Please create it first.")
                            continue
                            
                    if "semester" in col and row[col["semester"]]:
                        semester = int(row[col["semester"]])
                        # Guard: reject semesters that exceed the program's total
                        if program and semester > program.total_semesters:
                            errors.append(
                                f"Row {row_num}: semester {semester} exceeds "
                                f"{program.code}'s max ({program.total_semesters} sems). "
                                f"Course '{code}' skipped."
                            )
                            continue
                        
                    if "parent_pe_code" in col and row[col["parent_pe_code"]]:
                        parent_code = str(row[col["parent_pe_code"]]).strip()

                    obj, was_created = Course.objects.get_or_create(
                        code=code,
                        defaults={
                            "name": name,
                            "credits": credits_val,
                            "course_type": ct,
                            "program": program,
                            "semester": semester,
                            "parent_course_code": parent_code,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        # Ensure fields are updated
                        update_fields = []
                        if obj.name != name:
                            obj.name = name; update_fields.append("name")
                        if obj.credits != credits_val:
                            obj.credits = credits_val; update_fields.append("credits")
                        if obj.course_type != ct:
                            obj.course_type = ct; update_fields.append("course_type")
                        if obj.program != program:
                            obj.program = program; update_fields.append("program")
                        if obj.semester != semester:
                            obj.semester = semester; update_fields.append("semester")
                        if obj.parent_course_code != parent_code:
                            obj.parent_course_code = parent_code; update_fields.append("parent_course_code")

                        if update_fields:
                            obj.save(update_fields=update_fields)
                            updated += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

        return Response(
            {"created": created, "updated": updated, "errors": errors},
            status=status.HTTP_200_OK,
        )


# ----------------------------
# PROGRAM BULK UPLOAD
# ----------------------------

class ProgramBulkUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided. Upload an .xlsx file with field name 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith(".xlsx"):
            return Response(
                {"error": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            import openpyxl
        except ImportError:
            return Response(
                {"error": "openpyxl is not installed. Run: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"error": f"Could not read Excel file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response(
                {"error": "File must have a header row and at least one data row."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

        required = {"name", "code"}
        if not required.issubset(set(header)):
            return Response(
                {"error": f"Missing required columns: {required - set(header)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        col = {name: idx for idx, name in enumerate(header)}
        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    name = str(row[col["name"]] or "").strip()
                    code = str(row[col["code"]] or "").strip()

                    if not name or not code:
                        errors.append(f"Row {row_num}: name or code is empty, skipped.")
                        continue

                    # Resolve or auto-create department
                    department = None
                    if "department_code" in col and row[col["department_code"]]:
                        dept_code = str(row[col["department_code"]]).strip()
                        department, dept_created = Department.objects.get_or_create(
                            code__iexact=dept_code,
                            defaults={
                                "code": dept_code.upper(),
                                "name": f"Department of {dept_code.upper()}",
                            },
                        )
                        if dept_created:
                            log.info(f"Auto-created department '{dept_code}'")
                    elif "department" in col and row[col["department"]]:
                        dept_name = str(row[col["department"]]).strip()
                        department, dept_created = Department.objects.get_or_create(
                            name__iexact=dept_name,
                            defaults={
                                "name": dept_name,
                                "code": dept_name[:20].upper().replace(" ", ""),
                            },
                        )

                    if department is None:
                        # If no department info at all, create a default
                        department, _ = Department.objects.get_or_create(
                            code="GEN",
                            defaults={"name": "General"},
                        )

                    specialization = ""
                    if "specialization" in col and row[col["specialization"]]:
                        specialization = str(row[col["specialization"]]).strip()

                    total_years = 4
                    if "total_years" in col and row[col["total_years"]]:
                        total_years = int(row[col["total_years"]])

                    total_semesters = 8
                    if "total_semesters" in col and row[col["total_semesters"]]:
                        total_semesters = int(row[col["total_semesters"]])

                    obj, was_created = Program.objects.get_or_create(
                        code=code,
                        defaults={
                            "name": name,
                            "department": department,
                            "specialization": specialization,
                            "total_years": total_years,
                            "total_semesters": total_semesters,
                        },
                    )

                    if was_created:
                        created += 1
                    else:
                        obj.name = name
                        obj.department = department
                        obj.specialization = specialization
                        obj.total_years = total_years
                        obj.total_semesters = total_semesters
                        obj.save()
                        updated += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

        return Response(
            {"created": created, "updated": updated, "errors": errors},
            status=status.HTTP_200_OK,
        )