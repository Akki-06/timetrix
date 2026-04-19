# server/ml_pipeline/management/commands/seed_eligibility.py
"""
Parses the 'Eligibility Detail' sheet in Faculty.xlsx
and creates FacultySubjectEligibility records.

Sheet columns: faculty_name, course_code, course_name, program, semester, section, type
  type = "Theory" | "Lab" | "Theory+Lab"

Run AFTER faculty and courses are uploaded:
    python manage.py seed_eligibility
"""

import logging
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

log = logging.getLogger(__name__)

DATA = Path(__file__).resolve().parent.parent.parent.parent / "ml_pipeline" / "data"
FACULTY_XLSX = DATA / "Faculty.xlsx"


class Command(BaseCommand):
    help = "Create FacultySubjectEligibility records from Faculty Master Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--reset", action="store_true",
            help="Delete all existing eligibility records first",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        from faculty.models import Faculty, FacultySubjectEligibility
        from academics.models import Course

        if options["reset"]:
            count = FacultySubjectEligibility.objects.all().delete()[0]
            self.stderr.write(self.style.WARNING(
                f"Deleted {count} existing eligibility records."
            ))

        import openpyxl
        wb = openpyxl.load_workbook(FACULTY_XLSX, read_only=True)

        # Use 'Eligibility Detail' sheet
        sheet_name = "Eligibility Detail"
        if sheet_name not in wb.sheetnames:
            self.stderr.write(self.style.ERROR(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            ))
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.stderr.write(self.style.ERROR("Sheet is empty."))
            return

        # Parse header
        header = [str(c or "").strip().lower().replace(" ", "_") for c in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}

        required = {"faculty_name", "course_code"}
        missing = required - set(col)
        if missing:
            self.stderr.write(self.style.ERROR(
                f"Missing columns: {missing}. Found: {header}"
            ))
            return

        # Build lookups
        fac_by_name = {f.name.strip().lower(): f for f in Faculty.objects.filter(is_active=True)}

        # Course lookup by code (exact) and by code without program suffix
        course_by_code = {}
        for c in Course.objects.all():
            course_by_code[c.code.strip()] = c
            # Also index stripped version: "24COA274_BCA" → "24COA274"
            if "_" in c.code:
                stripped = c.code.rsplit("_", 1)[0]
                course_by_code.setdefault(stripped, c)

        created = 0
        already_existed = 0
        skipped_no_fac = 0
        skipped_no_course = 0

        for row in rows[1:]:
            fac_name = str(row[col["faculty_name"]] or "").strip()
            course_code = str(row[col["course_code"]] or "").strip()

            if not fac_name or not course_code:
                continue

            fac = fac_by_name.get(fac_name.lower())
            if not fac:
                skipped_no_fac += 1
                continue

            # Resolve course: try exact, then strip suffix
            course = course_by_code.get(course_code)
            if not course and "_" in course_code:
                stripped = course_code.rsplit("_", 1)[0]
                course = course_by_code.get(stripped)

            if not course:
                skipped_no_course += 1
                continue

            # Priority from type column
            type_val = ""
            if "type" in col:
                type_val = str(row[col["type"]] or "").strip().lower()
            priority = 3 if ("theory" in type_val and "lab" in type_val) or "theory+lab" in type_val \
                       else 2 if "theory" in type_val \
                       else 1  # lab or unknown

            _, was_created = FacultySubjectEligibility.objects.get_or_create(
                faculty=fac,
                course=course,
                defaults={"priority_weight": priority},
            )
            if was_created:
                created += 1
            else:
                already_existed += 1

        self.stdout.write(self.style.SUCCESS(
            f"Eligibility seeded: {created} created, {already_existed} already existed."
        ))
        if skipped_no_fac:
            self.stdout.write(self.style.WARNING(
                f"  {skipped_no_fac} rows skipped (faculty not found - upload faculty first)"
            ))
        if skipped_no_course:
            self.stdout.write(self.style.WARNING(
                f"  {skipped_no_course} rows skipped (course code not found in DB)"
            ))
