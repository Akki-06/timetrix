import logging

from django.db import transaction
from rest_framework import viewsets, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django_filters.rest_framework import DjangoFilterBackend

from .models import (
    Faculty, TeacherAvailability, FacultySubjectEligibility,
    FacultyProgramExclusion, FacultySemesterExclusion,
)
from .serializers import (
    FacultySerializer,
    TeacherAvailabilitySerializer,
    FacultySubjectEligibilitySerializer,
    FacultyProgramExclusionSerializer,
    FacultySemesterExclusionSerializer,
)

log = logging.getLogger(__name__)


# ----------------------------
# FACULTY
# ----------------------------

class FacultyViewSet(viewsets.ModelViewSet):
    queryset = Faculty.objects.all()
    serializer_class = FacultySerializer

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["role", "is_active"]
    search_fields = ["name", "employee_id"]
    ordering_fields = ["name", "employee_id", "role"]


# ----------------------------
# TEACHER AVAILABILITY
# ----------------------------

class TeacherAvailabilityViewSet(viewsets.ModelViewSet):
    queryset = TeacherAvailability.objects.select_related("faculty")
    serializer_class = TeacherAvailabilitySerializer

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["faculty", "day"]


# ----------------------------
# FACULTY SUBJECT ELIGIBILITY
# ----------------------------

class FacultySubjectEligibilityViewSet(viewsets.ModelViewSet):
    queryset = FacultySubjectEligibility.objects.select_related(
        "faculty",
        "course"
    )
    serializer_class = FacultySubjectEligibilitySerializer

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["faculty", "course"]


# ----------------------------
# FACULTY BULK UPLOAD
# ----------------------------

ROLE_MAP = {
    "pvc":                  "PVC",
    "pro vice chancellor":  "PVC",
    "dean":                 "DEAN",
    "dean of school":       "DEAN",
    "hod":                  "HOD",
    "head of department":   "HOD",
    "regular":              "REGULAR",
    "assistant professor":  "REGULAR",
    "associate professor":  "REGULAR",
    "professor":            "REGULAR",
    "senior":               "REGULAR",
    "senior faculty":       "REGULAR",
    "visiting":             "VISITING",
    "visiting faculty":     "VISITING",
    "contractual":          "VISITING",
}


class FacultyBulkUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided. Upload an .xlsx file with field name 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith(".xlsx"):
            return Response(
                {"error": "Only .xlsx files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            import openpyxl
        except ImportError:
            return Response(
                {"error": "openpyxl is not installed. Run: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {"error": f"Could not read Excel file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response(
                {"error": "File must have a header row and at least one data row."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

        required = {"name", "employee_id"}
        if not required.issubset(set(header)):
            return Response(
                {"error": f"Missing required columns: {required - set(header)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        col = {name: idx for idx, name in enumerate(header)}
        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            from academics.models import Department
            dept_cache = {}
            for d in Department.objects.all():
                dept_cache[d.name.strip().lower()] = d

            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    name = str(row[col["name"]] or "").strip()
                    emp_id = str(row[col["employee_id"]] or "").strip()

                    if not name or not emp_id:
                        errors.append(f"Row {row_num}: name or employee_id is empty, skipped.")
                        continue

                    role_raw = str(row[col["role"]] or "REGULAR").strip().lower() if "role" in col else "regular"
                    role = ROLE_MAP.get(role_raw, "REGULAR")

                    designation = str(row[col["designation"]] or "").strip() if "designation" in col else ""

                    max_weekly = int(row[col["max_weekly_load"]]) if "max_weekly_load" in col and row[col["max_weekly_load"]] else 18
                    max_daily = int(row[col["max_lectures_per_day"]]) if "max_lectures_per_day" in col and row[col["max_lectures_per_day"]] else 4
                    max_consec = int(row[col["max_consecutive_lectures"]]) if "max_consecutive_lectures" in col and row[col["max_consecutive_lectures"]] else 2

                    teaches_theory = True
                    if "teaches_theory" in col and row[col["teaches_theory"]] is not None:
                        val = str(row[col["teaches_theory"]]).strip().lower()
                        teaches_theory = val in ("1", "true", "yes")

                    teaches_lab = True
                    if "teaches_lab" in col and row[col["teaches_lab"]] is not None:
                        val = str(row[col["teaches_lab"]]).strip().lower()
                        teaches_lab = val in ("1", "true", "yes")

                    dept = None
                    if "department" in col and row[col["department"]]:
                        dept_name = str(row[col["department"]]).strip()
                        dept = dept_cache.get(dept_name.lower())

                    defaults = {
                        "name": name,
                        "role": role,
                        "designation": designation,
                        "max_weekly_load": max_weekly,
                        "max_lectures_per_day": max_daily,
                        "max_consecutive_lectures": max_consec,
                        "teaches_theory": teaches_theory,
                        "teaches_lab": teaches_lab,
                    }
                    if dept:
                        defaults["department"] = dept

                    obj, was_created = Faculty.objects.get_or_create(
                        employee_id=emp_id,
                        defaults=defaults,
                    )

                    if was_created:
                        created += 1
                    else:
                        obj.name = name
                        obj.role = role
                        obj.designation = designation
                        obj.max_weekly_load = max_weekly
                        obj.max_lectures_per_day = max_daily
                        obj.max_consecutive_lectures = max_consec
                        obj.teaches_theory = teaches_theory
                        obj.teaches_lab = teaches_lab
                        if dept:
                            obj.department = dept
                        obj.save()
                        updated += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

        return Response(
            {"created": created, "updated": updated, "errors": errors},
            status=status.HTTP_200_OK,
        )


# ----------------------------
# ELIGIBILITY BULK UPLOAD
# ----------------------------

class EligibilityBulkUploadView(APIView):
    """
    Upload an Excel with columns: faculty_name, course_code
    (optionally: priority_weight).
    Auto-matches faculty by name and course by code. Creates
    FacultySubjectEligibility records.
    """
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided."}, status=400)
        if not file.name.endswith(".xlsx"):
            return Response({"error": "Only .xlsx files supported."}, status=400)

        try:
            import openpyxl
        except ImportError:
            return Response({"error": "openpyxl not installed."}, status=500)

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"error": f"Cannot read file: {e}"}, status=400)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({"error": "Need header + data rows."}, status=400)

        header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}

        if "faculty_name" not in col or "course_code" not in col:
            return Response(
                {"error": "Required columns: faculty_name, course_code"},
                status=400,
            )

        from academics.models import Course

        # Build lookup caches
        fac_by_name = {}
        for f in Faculty.objects.filter(is_active=True):
            fac_by_name[f.name.strip().lower()] = f

        course_by_code = {}
        for c in Course.objects.all():
            course_by_code[c.code.strip().lower()] = c

        created = 0
        skipped = 0
        assigned = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    fac_name = str(row[col["faculty_name"]] or "").strip()
                    c_code   = str(row[col["course_code"]] or "").strip()

                    if not fac_name or not c_code:
                        errors.append(f"Row {row_num}: empty faculty_name or course_code.")
                        continue

                    fac = fac_by_name.get(fac_name.lower())
                    if not fac:
                        errors.append(f"Row {row_num}: faculty '{fac_name}' not found.")
                        continue

                    course = course_by_code.get(c_code.lower())
                    if not course:
                        errors.append(f"Row {row_num}: course '{c_code}' not found.")
                        continue

                    pw = 1
                    if "priority_weight" in col and row[col["priority_weight"]]:
                        pw = int(row[col["priority_weight"]])

                    _, was_created = FacultySubjectEligibility.objects.get_or_create(
                        faculty=fac,
                        course=course,
                        defaults={"priority_weight": pw},
                    )
                    if was_created:
                        created += 1
                    else:
                        skipped += 1

                    # Also directly assign faculty to all CourseOfferings for this course
                    from academics.models import CourseOffering
                    updated = CourseOffering.objects.filter(
                        course=course,
                        assigned_faculty__isnull=True,
                    ).update(assigned_faculty=fac)
                    assigned += updated

                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

        return Response(
            {"created": created, "skipped": skipped, "assigned": assigned, "errors": errors},
            status=status.HTTP_200_OK,
        )


# ----------------------------
# FACULTY WORKLOAD SUMMARY
# ----------------------------

class FacultyWorkloadView(APIView):
    """
    Returns each active faculty's max/current workload.
    current = number of FacultySubjectEligibility records
    (as proxy for assigned courses × lectures).
    """

    def get(self, request):
        from academics.models import CourseOffering
        from django.db.models import Count, Sum

        faculty_list = Faculty.objects.filter(is_active=True).order_by("name")
        # Count eligibility records per faculty
        elig_counts = dict(
            FacultySubjectEligibility.objects
            .values_list("faculty_id")
            .annotate(cnt=Count("id"))
            .values_list("faculty_id", "cnt")
        )
        # Count assigned offerings' weekly lectures
        assigned_load = dict(
            CourseOffering.objects
            .filter(assigned_faculty__isnull=False)
            .values("assigned_faculty_id")
            .annotate(total=Sum("course__min_weekly_lectures"))
            .values_list("assigned_faculty_id", "total")
        )

        data = []
        for f in faculty_list:
            current = assigned_load.get(f.id, 0) or 0
            data.append({
                "id": f.id,
                "name": f.name,
                "employee_id": f.employee_id,
                "role": f.role,
                "max_weekly_load": f.max_weekly_load,
                "current_load": current,
                "remaining": f.max_weekly_load - current,
                "eligible_courses": elig_counts.get(f.id, 0),
            })
        return Response(data)


# ----------------------------
# PROGRAM EXCLUSION
# ----------------------------

class FacultyProgramExclusionViewSet(viewsets.ModelViewSet):
    queryset = FacultyProgramExclusion.objects.all()
    serializer_class = FacultyProgramExclusionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["faculty", "program"]


# ----------------------------
# SEMESTER EXCLUSION
# ----------------------------

class FacultySemesterExclusionViewSet(viewsets.ModelViewSet):
    queryset = FacultySemesterExclusion.objects.all()
    serializer_class = FacultySemesterExclusionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["faculty", "program", "semester"]